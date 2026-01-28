#app.api.routes.py
from __future__ import annotations

from fastapi import APIRouter, HTTPException, Body, UploadFile, File
from fastapi.responses import PlainTextResponse
import io
import csv
from datetime import datetime
from typing import List
import uuid

from app.engine.generator import generate_insights
from app.api.schemas import GenerateInsightsRequest, GenerateInsightsResponse, ErrorResponse
from app.api.schemas import InsightType
from pathlib import Path

router = APIRouter(prefix="/v1/insights", tags=["insights"])


@router.post(
    "/generate",
    response_model=GenerateInsightsResponse,
    responses={400: {"model": ErrorResponse}, 500: {"model": ErrorResponse}},
)
def generate(req: GenerateInsightsRequest) -> GenerateInsightsResponse:
    try:
        return generate_insights(req)
    except ValueError as e:
        raise HTTPException(status_code=400, detail={"error": "bad_request", "details": str(e)})
    except Exception as e:
        raise HTTPException(status_code=500, detail={"error": "internal_error", "details": str(e)})





@router.post(
    "/bulk-generate",
    response_class=PlainTextResponse,
    responses={400: {"model": ErrorResponse}, 500: {"model": ErrorResponse}},
)
def bulk_generate(file: UploadFile = File(...)) -> PlainTextResponse:
    """Accept a CSV or XLSX file with columns: customer_id, current_balance, last_activity_date

    For each row where `current_balance` == 0 and `last_activity_date` is older than 180 days,
    build a minimal `GenerateInsightsRequest` and call `generate_insights` (v1) to produce suggestions.
    """
    # support CSV directly; XLSX requires openpyxl (optional)
    fname = file.filename or "upload"
    content = file.file.read()

    rows = []
    if fname.lower().endswith(".csv"):
        s = content.decode(errors="replace")
        reader = csv.DictReader(io.StringIO(s))
        for r in reader:
            rows.append(r)
    elif fname.lower().endswith(('.xls', '.xlsx')):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({h: v for h, v in zip(headers, row)})
        except Exception as e:
            raise HTTPException(status_code=400, detail={
                "error": "xlsx_not_supported",
                "details": "XLSX requires openpyxl installed on the server or upload CSV instead. " + str(e),
            })
    else:
        raise HTTPException(status_code=400, detail={"error": "unsupported_file", "details": "Provide CSV or XLSX file."})

    lines: List[str] = []
    now_iso = datetime.utcnow().isoformat() + "Z"

    for r in rows:
        # normalize column access (lowercase keys)
        row = {k.strip().lower(): v for k, v in (r.items() if isinstance(r, dict) else [])}
        try:
            balance = float(row.get('current_balance') or row.get('balance') or 0)
        except Exception:
            balance = 0

        last_activity = row.get('last_activity_date') or row.get('last_activity') or row.get('last_login_at')
        last_dt = None
        if last_activity:
            try:
                last_dt = datetime.fromisoformat(str(last_activity))
            except Exception:
                try:
                    last_dt = datetime.strptime(str(last_activity), "%Y-%m-%d")
                except Exception:
                    last_dt = None

        # check criteria: zero balance and inactive >180 days
        inactive = False
        if last_dt:
            inactive = (datetime.utcnow() - last_dt).days > 180

        if balance == 0 and inactive:
            # Use a static curated tickers file to produce top-3 suggestions per matching row.
            try:
                static_file = Path("app/data/static_top_tickers.txt")
                cid = row.get('customer_id') or 'unknown'
                if not static_file.exists():
                    lines.append(f"{cid}: (no static list available)")
                    continue

                with static_file.open("r", encoding="utf-8") as fh:
                    entries = []
                    for line in fh:
                        line = line.strip()
                        if not line:
                            continue
                        parts = [p.strip() for p in line.split("|")]
                        symbol = parts[0].upper()
                        name = parts[1] if len(parts) > 1 and parts[1] else symbol
                        pct = parts[2] if len(parts) > 2 else ""
                        note = parts[3] if len(parts) > 3 else "strong recent momentum"
                        entries.append({"symbol": symbol, "name": name, "pct": pct, "note": note})

                top_n = entries[:3]
                if not top_n:
                    lines.append(f"{cid}: (no candidates)")
                    continue

                # Build human-friendly recommendation header including inactivity length
                if last_dt:
                    months = max(1, (datetime.utcnow() - last_dt).days // 30)
                    month_label = "months" if months > 1 else "month"
                    header = (
                        f"We noticed you haven't been active for about {months} {month_label}. "
                        "Here are three stocks that have been performing well recently to help you get back into the flow.\n"
                    )
                else:
                    header = (
                        "Welcome back — here are three stocks that have been performing well recently to help you get back into the flow.\n"
                    )

                # Short, conversational ticker lines
                ticker_lines = []
                for e in top_n:
                    pct_part = f" {e['pct']}" if e['pct'] else ""
                    # Example: AAPL (Apple Inc.) — strong recent momentum (+3.5%)
                    ticker_lines.append(f"{e['symbol']} ({e['name']}) — {e['note']}{pct_part}")

                # Add a brief CTA per user's request (no informational disclaimer)
                footer = "\nIf you'd like, we can add these to a watchlist or send a short market note. If you need any help from our end, we can help you engage them. That can help you to reach your goals."

                block = header + "\n".join(ticker_lines) + footer
                lines.append(f"{cid}: {block}")
            except Exception:
                lines.append(f"{row.get('customer_id') or 'unknown'}: (error generating suggestion)")
                continue

    if not lines:
        return PlainTextResponse(content="No matching rows found or no suggestions generated.")
    return PlainTextResponse(content="\n\n".join(lines))

